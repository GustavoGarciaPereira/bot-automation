"""Export cross-reference report to Excel with multiple sheets."""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)


class CrossReportExporter:
    """Exports cross-reference report to Excel."""

    def export(self, report: dict, output_dir: str = "data/output/cross_reference") -> str:
        try:
            import openpyxl
        except ImportError:
            logger.error("openpyxl not installed")
            return ""

        Path(output_dir).mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y-%m-%d_%H%M")
        filepath = f"{output_dir}/relatorio_360_{ts}.xlsx"

        wb = openpyxl.Workbook()

        # Sheet 1: Summary
        ws = wb.active
        ws.title = "Resumo"
        ws.append(["Campo", "Valor"])
        ws.append(["Gerado em", report.get("generated_at", "")])
        for k, v in report.get("summary", {}).items():
            ws.append([k, v])
        ws.append([])
        ws.append(["Insights"])
        for ins in report.get("insights", []):
            ws.append(["", ins])

        # Sheet 2: Price comparison
        ws2 = wb.create_sheet("Precos")
        comps = report.get("price_comparisons", [])
        if comps:
            headers = list(comps[0].keys())
            ws2.append(headers)
            for c in comps:
                ws2.append([c.get(h, "") for h in headers])

        # Sheet 3: Leads
        ws3 = wb.create_sheet("Leads")
        leads = report.get("enriched_leads", [])
        if leads:
            cols = [
                "name", "category", "address", "phone", "website",
                "rating", "trust_score", "trust_level", "insight",
                "online_presence_ml", "online_presence_olx",
            ]
            ws3.append(cols)
            for lead in leads:
                ws3.append([lead.get(c, "") for c in cols])

        wb.save(filepath)
        logger.info("Cross-report saved: %s", filepath)
        return filepath
