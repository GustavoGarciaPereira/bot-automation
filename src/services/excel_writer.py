"""Excel workbook writer using pandas + openpyxl.

Produces a formatted ``.xlsx`` file at ``data/output/<client_id>/``.
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import TYPE_CHECKING

import pandas as pd

from src.interfaces.output_writer import OutputWriter
from src.utils.logger import get_logger

if TYPE_CHECKING:
    from src.models import IntimacaoRecord

logger = get_logger(__name__)

# Column mapping: model_field → display_name
COLUMN_MAP: dict[str, str] = {
    "data_consulta": "collected_at",
    "portal": "platform",
    "advogado": "bot",
    "sequencia": "seq",
    "tipo_comunicacao": "type",
    "objeto_comunicacao": "title",
    "destinatario": "location",
    "instancia": "category",
    "comarca": "seller",
    "parte_1": "extra",
    "status_registro": "status",
    "despacho": "detail",
    "classification_confidence": "ai_score",
}

# Columns to show (in this order), using display names
DISPLAY_COLUMNS = [
    "collected_at",
    "platform",
    "bot",
    "type",
    "title",
    "category",
    "seller",
    "location",
    "status",
    "detail",
    "ai_score",
]


class ExcelWriter(OutputWriter):
    """Write records to an Excel workbook with auto-column-width and
    frozen header row.
    """

    async def write_records(
        self, records: list[IntimacaoRecord], client_id: str, data_ref: str
    ) -> str:
        """Persist *records* and return the absolute output path."""
        if not records:
            logger.info("No records to write for client %r — skipping Excel", client_id)
            return ""

        # ---- Build DataFrame -------------------------------------------
        df = pd.DataFrame([r.model_dump() for r in records])

        # Rename columns to display names
        df.rename(columns=COLUMN_MAP, inplace=True, errors="ignore")

        # Keep only display columns that exist in the data
        final_columns = [c for c in DISPLAY_COLUMNS if c in df.columns]
        df = df[final_columns]

        # ---- Output path ------------------------------------------------
        output_dir = Path("data/output") / client_id
        output_dir.mkdir(parents=True, exist_ok=True)

        filepath = output_dir / f"relatorio_{client_id}_{data_ref}.xlsx"

        # ---- Write with formatting --------------------------------------
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Relatório")

            ws = writer.sheets["Relatório"]

            # Auto-fit column widths
            for col_idx, col_name in enumerate(final_columns, start=1):
                max_len = max(
                    len(str(col_name)),
                    df[col_name].astype(str).str.len().max() if not df.empty else 0,
                )
                ws.column_dimensions[
                    ws.cell(row=1, column=col_idx).column_letter
                ].width = min(max_len + 3, 60)

            # Freeze header row
            ws.freeze_panes = "A2"

            # Header style
            from openpyxl.styles import Alignment, Font, PatternFill

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="2F5496", end_color="2F5496", fill_type="solid"
            )
            header_align = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

        logger.info(
            "Excel written: %s (%d rows × %d cols)",
            filepath,
            len(df),
            len(final_columns),
        )
        return str(filepath.resolve())
